#!usr/bin/python3
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(filename = "student.xlsx")
ws = wb["Sheet1"]

row = 2

while True:
	midterm = ws.cell(row = row, column = 3).value
	final = ws.cell(row = row, column = 4).value
	hw = ws.cell(row = row, column = 5).value
	attendance = ws.cell(row = row, column = 6).value

	if midterm == None:
		break
	
	total = midterm * 0.3 + final * 0.35 + hw * 0.34 + attendance

	grade = 'Q'

	ws.cell(row = row, column = 7, value = total)
	ws.cell(row = row, column = 8, value = grade)
	row += 1

stuNum = row - 2
AP = int(stuNum * 0.15)
A = int(stuNum * 0.3) - AP
BP = int(stuNum * 0.2)
B = int(stuNum * 0.4) - BP
re = stuNum - AP - A - BP - B 
CP = re // 2
C = re - CP

gradeNum = [AP, A, BP, B, CP, C]
print(gradeNum)
gradeStr = ["A+", "A0", "B+", "B0", "C+", "C0"]

num = 0
n = 0

while num != stuNum:
	best = -100
	index = -1
	i = -1
	for i in range (2, row):
		if best < ws.cell(row = i, column = 7).value:
			if ws.cell(row = i, column = 8).value != 'Q':
				continue
			best = ws.cell(row = i, column = 7).value
			best_index = i
	if gradeNum[n] == 0:
		n += 1
		continue
	if ws.cell(row = best_index, column = 8).value == 'Q':
		ws.cell(row = best_index, column = 8, value = gradeStr[n])
		gradeNum[n] -= 1
		num += 1
        
wb.save(filename = "student.xlsx")

